import  os.path
import shutil
import zipfile
from io import BytesIO
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook

import pytest

@pytest.fixture(scope="session")
def create_zip():
    source_dir = 'example_files'
    output_dir = 'archive'
    archive_name = 'test_archive.zip'
    archive_path = os.path.join(output_dir, archive_name)

    with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(source_dir):
            for file in files:
                file_path = os.path.join(root, file)
                zipf.write(file_path, os.path.relpath(file_path, source_dir))

    print(f'Архив создан: {archive_path}')

def test_pdf(create_zip):
    with ZipFile("archive\\test_archive.zip") as zip_file:
        reader = PdfReader(zip_file.open('sample3.pdf'))
        assert "Get Lost!" in reader.pages[0].extract_text()

def test_xlsx(create_zip):
    with ZipFile("archive\\test_archive.zip") as zip_file:
       workbook = load_workbook(zip_file.open('examp1.xlsx'))
       sheet = workbook.active
       assert sheet.cell(row=2, column=8).value == "1-1-3-1"

def test_csv(create_zip):
    with ZipFile("archive\\test_archive.zip") as zip_file:
        csv = zip_file.read('example.csv')
        assert str(csv) == "b'Name,Job Title,Address,State,City\\nJohn Doe,Designer,325 Pine Street,,Seattle\\n,,,,\\nEdward Green,Developer,110 Pike Street,WA,Seattle\\n'"
